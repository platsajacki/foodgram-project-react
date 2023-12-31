from io import BytesIO
from http import HTTPStatus
from typing import Any

from django.db.models import QuerySet
from django.http.response import HttpResponse
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework.views import exception_handler

from openpyxl_styles import header_style, body_style


def custom_exception_handler(
        exc: HttpResponse, context: dict[str, Any]
) -> Response:
    """
    Кастомный обработчик исключений для Django REST Framework.
    """
    response: Response = exception_handler(exc, context)
    if response and response.status_code == HTTPStatus.NOT_FOUND:
        response.data: dict[str, str] = {
            'detail': 'Страницы, которую Вы ищете, не существует.'
        }
    return response


def get_xls_shopping_cart(ingredients: QuerySet) -> BytesIO:
    """
    Функция для создания файла Excel (XLS)
    со списком покупок на основе переданных ингредиентов.
    """
    wb: Workbook = Workbook()
    wb.add_named_style(header_style)
    wb.add_named_style(body_style)
    ws = wb.active
    ws.column_dimensions['A'].width = 4
    ws.column_dimensions['B'].width = 65
    ws.column_dimensions['C'].width = 20
    ws.append(
        ['№', 'Ингредиент', 'Количество']
    )
    for cell in ws[1]:
        cell.style = 'header_style'
    number: int = 1
    for ingredient in ingredients:
        name: str = ingredient['recipe__recipeingredient__ingredient__name']
        total_amount: str = ingredient['total_amount']
        unit: str = ingredient[
            "recipe__recipeingredient__ingredient__measurement_unit"
        ]
        ws.append([number, name, f'{total_amount} {unit}'])
        for cell in ws[1 + number]:
            cell.style = 'body_style'
        number += 1
    buffer: BytesIO = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
